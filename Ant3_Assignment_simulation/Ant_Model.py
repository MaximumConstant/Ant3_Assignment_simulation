import pygame
from pygame.locals import *
import sys
import random as rnd
import math
from PIL import Image, ImageDraw,ImageFont
import numpy as np
import time
import openpyxl
from munkres import Munkres

def OnTrail(agent,trail1,stage,S):
    next_point=trail1[stage]
    dist_among=math.hypot(agent_x[agent]-point_x[next_point],agent_y[agent]-point_y[next_point])
    if(dist_among<S):
        agent_x[agent]=point_x[next_point]
        agent_y[agent]=point_y[next_point]
        if(stage==len(trail1)-1):
            endflag=1
            return endflag,stage
        else:
            return OnTrail(agent,trail1,stage+1,S-dist_among)
    else:
        alpha=math.atan2(point_y[next_point]-agent_y[agent],point_x[next_point]-agent_x[agent])
        agent_x[agent]+=S*math.cos(alpha)
        agent_y[agent]+=S*math.sin(alpha)
        endflag=0
        return endflag,stage

def DisplayAll():
    for i in pygame.event.get():
            if i.type == pygame.QUIT:
                sys.exit()
    sc.fill(WHITE)
    sc.blit(bg,(0,0))
    for i in range(n_agent):
        if(endflag[i]==0):
            pygame.draw.circle(sc, ORANGE,(agent_x[i], agent_y[i]), Rob_gabarite)
        else:
            pygame.draw.circle(sc, BLACK,(agent_x[i], agent_y[i]), Rob_gabarite)
    for i in range(nCity):
      pygame.draw.circle(sc, BLACK,(point_x[i], point_y[i]), 1)
    pygame.display.update()
   
def image_to_points(Blocks_picture):
  img = Image.open(Blocks_picture)
  img = img.convert('1')
  first_arr = np.asarray(img, dtype='int')
  x_shape=first_arr.shape[1]
  y_shape=first_arr.shape[0]
  arr=[[0 for j in range(y_shape)] for i in range(x_shape)]
  arr2=[[0 for j in range(y_shape)] for i in range(x_shape)]
  for x_arr in range(x_shape):
    for y_arr in range(y_shape):
      arr[x_arr][y_arr]=(math.ceil((first_arr[y_arr][x_arr]/255)))*(-1)+1
      if (arr[x_arr][y_arr]==1):
        for x in range(Rob_gabarite+1):
          for y in range(x):
            if(x_arr+x-1>0 and y_arr+y-1<y_shape-1):
              arr2[x_arr+x-1][y_arr+y-1]=1
            if(x_arr+x-1>0 and y_arr-y-1>0):
              arr2[x_arr+x-1][y_arr-y-1]=1
            if(x_arr-x-1<x_shape-1 and y_arr-y-1>0):
              arr2[x_arr-x-1][y_arr-y-1]=1
            if(x_arr-x-1<x_shape-1 and y_arr+y-1<y_shape-1):
              arr2[x_arr-x-1][y_arr+y-1]=1
  return x_shape,y_shape,arr,arr2

def do_point_map_cage_im(): #Создания сетки на плоскости
  space=math.floor((((near_range**2)/2)**0.5)-2)
  x_len=math.floor(window_x/space)
  y_len=math.floor(window_y/space)
  nCity=x_len*y_len
  dist_among = [[0 for j in range(nCity)] for i in range(nCity)]
  schet=0
  for now_y in range(y_len):
    for now_x in range(x_len):
      if(arr2[math.ceil(now_x*space+space/2)][math.ceil(now_y*space+space/2)]==0):
        thisnear=[]
        point_x.append(math.ceil(now_x*space+space/2))
        point_y.append(math.ceil(now_y*space+space/2))
        now=schet
        for other in range(now):
          if ((((point_x[other]-point_x[now])**2 + (point_y[other]-point_y[now])**2)**0.5)<near_range):
            dist=((point_x[other]-point_x[now])**2 + (point_y[other]-point_y[now])**2)**0.5
            proof_space=math.ceil(dist/5)
            proof=1
            for n in range(proof_space):
              tx=math.ceil(point_x[now]+(point_x[other]-point_x[now])/proof_space*n)
              ty=math.ceil(point_y[now]+(point_y[other]-point_y[now])/proof_space*n)
              if(arr2[tx][ty]):
                proof=0
                break
            if(proof==1):
              thisnear.append(other)
              point_near[other].append(now)
              dist_among[now][other]=dist
              dist_among[other][now]=dist
        point_near.append(thisnear)
        schet=schet+1
  nCity=schet
  return nCity,dist_among

def DeikstrOnePoint(Now,NotBlock,MinRouteL,MinR):
  for Look in (point_near[Now]):
    if(dist_among[Now][Look]+MinRouteL[Now]<MinRouteL[Look]):
      MinRouteL[Look]=dist_among[Now][Look]+MinRouteL[Now]
      flag=0
      for i in (NotBlock):
        if(i==Look):
          flag=1
          break
      if (flag==0):
        NotBlock.append(Look)
      MinR[Look]=MinR[Now].copy()
      MinR[Look].append(Look)
  NotBlock.remove(Now)
  
def Deikstra(agent,end1):
  nearestPoint=0
  nearestLength=math.hypot(agent_x[agent]-point_x[nearestPoint],agent_y[agent]-point_y[nearestPoint])
  for i in range(nCity):
    NewTry=math.hypot(agent_x[agent]-point_x[i],agent_y[agent]-point_y[i])
    if(nearestLength>NewTry):
      nearestLength=NewTry
      nearestPoint=i
  start1=nearestPoint
  NotBlocked=[]
  NotBlocked.append(start1)
  MinRouteLen=[]
  for i in range(nCity):
    MinRouteLen.append(100000000)
  MinRoute=[]
  for j in range(nCity):
    OnePoint=[]
    MinRoute.append(OnePoint)
  MinRoute[start1].append(start1)
  MinRouteLen[start1]=0
  DeikstrOnePoint(start1,NotBlocked,MinRouteLen,MinRoute)
  while(1):
    if(len(NotBlocked)==0):
      break
    min=NotBlocked[0]
    for k in (NotBlocked):
      if(MinRouteLen[min]>MinRouteLen[k]):
        min=k
    DeikstrOnePoint(min,NotBlocked,MinRouteLen,MinRoute)
    if(min==end1):
      break
  if(MinRouteLen[end1]==100000000):
    print("No route!")
  return MinRoute[end1],MinRouteLen[end1]

def PutPher(Matches,l,deltaPher): #Распределение феромонов после движения муравьёв
  for i in range(len(Matches)):
    l1=l
    if(l1<1):
      l1=1
    deltaPher[i][Matches[i]]+=Q/l1

def UpdPher(deltaPher,pherTrail): # Обновление феромонов
  for i in range(n_agent):
    for j in range (n_agent):
      pherTrail[i][j]=pherTrail[i][j]*(1-Rho)+deltaPher[i][j]
      if(pherTrail[i][j]<pherMin):
        pherTrail[i][j]=pherMin
      if(pherTrail[i][j]>pherMax):
        pherTrail[i][j]=pherMax
      deltaPher[i][j]=0

def probability(agent,goal,pherTrail): #Расчёт весов определённого отрезка пути для расчёта вероятности прохода муравья по нему.
  l1=lenTrail[agent][goal]
  if(l1<1):
    l1=1
  p=(pherTrail[agent][goal]**Alpha)*((1/l1)**Beta)
  return p

def ChoiseGoal(agent,NotTabu,pherTrail): #Функция выбора следующей точки для перехода
  WhereList=[]
  ProbList=[]
  for cit in NotTabu:
    WhereList.append(cit)
    ProbList.append(probability(agent,cit,pherTrail))
  return rnd.choices(WhereList, weights=ProbList)[0]

def FindMatch(now,pherTrail): #Функция перехода муравья из одной точки в другую
  NotTabu=[]
  FreeAgent=[]
  NewMatches = [-1 for i in range(n_agent)]
  nextA=-1
  for i in range (n_agent):
    FreeAgent.append(i)
    NotTabu.append(i)
  while len(FreeAgent)>0:
    if(nextA==-1):
      nextA=now
    else:
      nextA=FreeAgent[rnd.randint(0,len(FreeAgent)-1)]
    Match=ChoiseGoal(nextA,NotTabu,pherTrail)
    NewMatches[nextA]=Match
    NotTabu.remove(Match)
    FreeAgent.remove(nextA)
  return NewMatches

def Length(Trail): #Расчёт верхней границы стоимости
  l=-1
  for i in range(len(Trail)):
    new_l=lenTrail[i][Trail[i]]
    if(new_l>l or l==-1):
      l=new_l
  return l

def SummLen(Trail):
  l=0
  for i in range(len(Trail)):
    l+=lenTrail[i][Trail[i]]
  return l

def AntAdmin():
  global bestLen,bestSumm
  pherTrail = [[pherMin for j in range(n_agent)] for i in range(n_agent)]
  deltaPher = [[0 for j in range(n_agent)] for i in range(n_agent)]
  bestLen=-1
  bestSumm=0
  bestTrail=[]
  BestList=[]
  BestSumm=[]
  StepList=[]
  bestTime=0
  step=0
  while (bestTime>step-200):
    for Ant in range(n_agent):
      way=FindMatch(Ant,pherTrail)
      l=Length(way)
      l2=SummLen(way)
      if(bestLen==-1 or l<bestLen):
        bestLen=l
        bestSumm=l2
        bestTrail=way
        bestTime=step
        BestList.append(bestLen)
        BestSumm.append(bestSumm)
        StepList.append(step)
        #print("На шаге",step, " найден путь длиной",bestLen)
      if(bestLen==l and l2<bestSumm):
        bestLen=l
        bestSumm=l2
        bestTrail=way
        bestTime=step
        BestList.append(bestLen)
        BestSumm.append(bestSumm)
        StepList.append(step)
      PutPher(way,l,deltaPher)
    UpdPher(deltaPher,pherTrail)
    step+=1
  makeExel(StepList,BestList,BestSumm)
  return bestTrail,bestLen,bestSumm

def makeExel(StepList,BestList,BestSumm):
  wb = openpyxl.Workbook()
  wb.create_sheet(title = 'Ant', index = 0)
  sheet = wb['Ant']
  NowCell = sheet.cell(row = 1, column = 1)
  NowCell.value = "Step"
  NowCell = sheet.cell(row = 1, column = 2)
  NowCell.value = "FoundLength"
  NowCell = sheet.cell(row = 1, column = 3)
  NowCell.value = "Summ"
  for i in range (len(BestList)):
    NowCell = sheet.cell(row = i+2, column = 1)
    NowCell.value = str(StepList[i])
    NowCell = sheet.cell(row = i+2, column = 2)
    NowCell.value = str(BestList[i])
    NowCell = sheet.cell(row = i+2, column = 3)
    NowCell.value = str(BestSumm[i])
  wb.save('AntMatches.xlsx')

def Greed():
  NotBlock=[]
  Matches=[-1 for j in range(n_agent)]
  for i in range(n_agent):
    NotBlock.append(i)
  for i in range(n_agent):
    flag=0
    l_min=-1
    for j in NotBlock:
      if(flag==0 or lenTrail[i][j]<lenTrail[i][n_min]):
        n_min=j
        flag=1
    Matches[i]=n_min
    NotBlock.remove(n_min)
  print('Жадный. Максимальная: ',Length(Matches), " Сумма: ",SummLen(Matches))
  return Matches

def GrossOne(step,NotBlock,Matches):
  global GrossMin_Match,GrossMin_l,Gross_summ
  if(len(NotBlock)==0):
      if (Length(Matches)<GrossMin_l):
        GrossMin_Match=Matches
        GrossMin_l=Length(Matches)
        Gross_summ=SummLen(Matches)
      if(Length(Matches)==GrossMin_l and SummLen(Matches)<Gross_summ):
        GrossMin_Match=Matches
        GrossMin_l=Length(Matches)
        Gross_summ=SummLen(Matches)
  else:
    for i in NotBlock:
      m1=Matches.copy()
      m1.append(i)
      n1=NotBlock.copy()
      n1.remove(i)
      GrossOne(step,n1,m1)

def Gross():
  sys.setrecursionlimit(1000000)
  global GrossMin_Match,GrossMin_l,Gross_summ
  GrossNotBlock=[]
  GrossMatches=[]
  GrossMin_Match=[]
  GrossMin_l=100000
  Gross_summ=0
  for i in range(n_agent):
    GrossNotBlock.append(i)
  GrossOne(0,GrossNotBlock,GrossMatches)
  print('Перебор. Максимальная: ',Length(GrossMin_Match), " Сумма: ",SummLen(GrossMin_Match))
  
    
def Hungarian():
  m = Munkres()
  indexes = m.compute(lenTrail)
  min_l = 0
  summ_l=0
  for row, column in indexes:
    value = lenTrail[row][column]
    summ_l+=value
    if(min_l<value): 
      min_l = value
    #print(f'({row}, {column}) -> {value}')
  print('Венгерский. Максимальная: ',min_l, " Сумма: ",summ_l)


FPS = 30
delta_time=1/FPS

WHITE = (255, 255, 255)
ORANGE = (255, 150, 100)
BLACK = (0, 0, 0)
RED = (255, 0, 0)
BLUE = (0,0, 255)
GREEN = (0, 170, 0)

n_agent=5
speed=30
near_range=30
Rob_gabarite=5

point_x=[]
point_y=[]
point_near=[]
agent_x=[]
agent_y=[]
stage=[]
trail=[]
lenTrail=[]
endflag=[]

Tend=400 #Время симуляции
Alpha=1 #Коэффициент альфа (порядка значимости феромона)
Beta=2 #Коэффициент бэта (порядка значимости длины пути)
Rho=0.2 #Коэффициент испарения феромона
Q=20 #Коэффициент увеличения феромона
pherMin=1 #Минимальное количество феромона на рёбрах графа
pherMax=5 #Максимальное количество феромона на рёбрах графа

goals=[0,1,2,4,6,21,23,25,27,42,44,47,59,61,63]
#goals=[0,1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40]

print("Stage 1: Loading of image")
adress="C:\Python\space.png"
bg = pygame.image.load(adress)
print("Stage 2: Analysis of image")
window_x,window_y,arr,arr2=image_to_points(adress)
print("Stage 3: Making of map")
nCity,dist_among=do_point_map_cage_im()

for i in range(n_agent):
   while(1):
      new_x=rnd.randint(1,window_x-2)
      new_y=rnd.randint(1,window_y-2)
      if(arr2[math.ceil(new_x)][math.ceil(new_y)]==0):
         break
   agent_x.append(new_x)
   agent_y.append(new_y)
   stage.append(0)
   endflag.append(0)
   trail.append([])
   lenTrail.append([])

clock = pygame.time.Clock()
sc = pygame.display.set_mode((window_x, window_y))
name="AntTask: n="+str(n_agent)
pygame.display.set_caption(name)
t1=time.process_time()
time_end=0

for i in range(n_agent):
   for j in range(n_agent):
      t,lt=Deikstra(i,goals[j])
      trail[i].append(t)
      lenTrail[i].append(lt)
      
t2=time.process_time()
time_dif=t2-t1
time_end=time_dif
if (time_dif<1):
  time_dif=2
circle=1
while 1:
   print("\nCircle = ",circle)
   circle+=1
   t1=time.process_time()
   for i in range(n_agent):
      for j in range(n_agent):
         trail[i][j],lenTrail[i][j]=Deikstra(i,goals[j])
      stage[i]=0
      endflag[i]=0
   t2=time.process_time()
   Matches,BestLenAnt,BestSumm=AntAdmin()
   print("Муравей. Максимальная: ", BestLenAnt," Сумма: ",BestSumm)
   t3=time.process_time()
   #print("Муравей: ")
   #for i in range(n_agent):
    #print(i," : ",Matches[i])
   GreedMatches=Greed()
   t4=time.process_time()
   Gross()
   t5=time.process_time()
   Hungarian()
   t6=time.process_time()
   time_dif=t3-t1
   now_time=0
   time_end+=time_dif
   if (time_dif<1.5):
    time_dif=1.5
   print("TimeDei = ",t2-t1)
   print("TimeAnt = ",t3-t2)
   print("TimeGreed = ",t4-t3)
   print("TimeGross = ",t5-t4)
   print("TimeHung = ",t6-t5)
   while now_time<time_dif:
      for i in range(n_agent):
         if(endflag[i]==0):
            S_1=speed*delta_time
            endflag[i],stage[i]=OnTrail(i,trail[i][Matches[i]],stage[i],S_1)
      DisplayAll()
      pygame.draw.rect(sc,GREEN,(0, 0, window_x, window_x), 2)  
      pygame.display.update()
      clock.tick(FPS)
      now_time+=delta_time
   pygame.draw.rect(sc,RED,(0, 0, window_x, window_x), 2)  
   pygame.display.update()